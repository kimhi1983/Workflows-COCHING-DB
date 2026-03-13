#!/usr/bin/env python3
"""
COCHING 데이터 Excel 누적 저장 스크립트
- 가이드처방 → formulations.xlsx (시트: 처방목록, 배합표상세)
- DB 백업 → db_ingredients.xlsx (시트: 원료, 규제, 지식베이스)
- n8n 수집 → collection_log.xlsx (시트: 수집이력)
- 30분마다 cron 실행, 누적 방식 (기존 데이터 유지 + 신규 추가)
"""
import json, os, re, glob
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl 미설치. pip install openpyxl")
    exit(1)

# ─── 경로 설정 ────────────────────────────────────────────────────────────────
BACKUP_BASE = Path("/mnt/e/COCHING-WORKFLOW/backup")
EXCEL_DIR = BACKUP_BASE / "excel"
FORMULATIONS_DIR = BACKUP_BASE / "formulations"
DB_DIR = BACKUP_BASE / "db"

EXCEL_DIR.mkdir(parents=True, exist_ok=True)

# ─── 스타일 설정 ──────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DATA_FONT = Font(name="맑은 고딕", size=10)
TITLE_FONT = Font(name="맑은 고딕", bold=True, size=13, color="2F5496")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")


def style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_row(ws, row, col_count, is_alt=False):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if is_alt:
            cell.fill = ALT_FILL


def auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                val_len = len(str(cell.value))
                if val_len > max_len:
                    max_len = val_len
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


# ═══════════════════════════════════════════════════════════════════════════════
# 1. 가이드처방 → formulations.xlsx
# ═══════════════════════════════════════════════════════════════════════════════
def export_formulations():
    excel_path = EXCEL_DIR / "COCHING_가이드처방.xlsx"

    # 기존 데이터 로드 (누적)
    existing_keys = set()
    if excel_path.exists():
        wb = load_workbook(excel_path)
        if "처방목록" in wb.sheetnames:
            ws = wb["처방목록"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    existing_keys.add(str(row[0]))  # 타임스탬프 기준
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # 처방목록 시트
    if "처방목록" not in wb.sheetnames:
        ws_list = wb.create_sheet("처방목록")
        headers = ["생성일시", "제품유형", "피부타입", "AI모델", "소요시간(초)",
                    "DB원료수", "규제수", "성분수", "파일명"]
        ws_list.append(headers)
        style_header(ws_list, 1, len(headers))
    else:
        ws_list = wb["처방목록"]

    # 배합표상세 시트
    if "배합표상세" not in wb.sheetnames:
        ws_detail = wb.create_sheet("배합표상세")
        headers = ["생성일시", "제품유형", "피부타입", "Phase", "순서",
                    "INCI명", "한글명", "배합비(%)", "기능", "비고"]
        ws_detail.append(headers)
        style_header(ws_detail, 1, len(headers))
    else:
        ws_detail = wb["배합표상세"]

    # JSON 파일 스캔
    new_count = 0
    for f in sorted(FORMULATIONS_DIR.glob("guide_*.json")):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            ts = data.get("timestamp", "")[:19]
            if ts in existing_keys:
                continue

            meta = data.get("metadata", {})
            formulation_text = data.get("formulation", "")

            # 성분 수 추출
            inci_matches = re.findall(r'"inci_name":\s*"([^"]+)"', formulation_text)

            row_idx = ws_list.max_row + 1
            ws_list.append([
                ts,
                data.get("product_type", ""),
                data.get("skin_type", ""),
                meta.get("model", ""),
                meta.get("elapsed", ""),
                meta.get("db_ingredients", ""),
                meta.get("regulations", ""),
                len(inci_matches),
                f.name,
            ])
            style_data_row(ws_list, row_idx, 9, is_alt=(new_count % 2 == 1))

            # 배합표 상세 추출
            # JSON 배열에서 formulation 배열 파싱 시도
            formulation_items = []
            try:
                # ```json ... ``` 블록 내부 추출
                json_blocks = re.findall(r'```json\s*([\s\S]*?)```', formulation_text)
                for block in json_blocks:
                    parsed = json.loads(block)
                    if isinstance(parsed, dict):
                        ft = parsed.get("2_formulation_table", {}).get("formulation", [])
                        if not ft:
                            ft = parsed.get("formulation", [])
                        if isinstance(ft, list):
                            formulation_items = ft
                            break
            except Exception:
                pass

            for item in formulation_items:
                detail_row = ws_detail.max_row + 1
                ws_detail.append([
                    ts,
                    data.get("product_type", ""),
                    data.get("skin_type", ""),
                    item.get("phase", ""),
                    item.get("order", ""),
                    item.get("inci_name", ""),
                    item.get("korean_name", ""),
                    item.get("percentage_wt", ""),
                    item.get("function", ""),
                    item.get("note", ""),
                ])
                style_data_row(ws_detail, detail_row, 10, is_alt=(detail_row % 2 == 0))

            new_count += 1
            existing_keys.add(ts)
        except Exception as e:
            print(f"  오류: {f.name} — {e}")

    auto_width(ws_list)
    auto_width(ws_detail)
    wb.save(excel_path)
    print(f"  가이드처방 Excel: {new_count}건 추가 (총 {ws_list.max_row - 1}건) → {excel_path.name}")


# ═══════════════════════════════════════════════════════════════════════════════
# 2. DB 백업 → db_master.xlsx
# ═══════════════════════════════════════════════════════════════════════════════
def export_db_data():
    excel_path = EXCEL_DIR / "COCHING_원료DB.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    # ─── 원료 마스터 시트 ─────────────────────────────────────────────────────
    latest_im = sorted(DB_DIR.glob("ingredient_master_*.json"), key=os.path.getmtime)
    if latest_im:
        ws = wb.create_sheet("원료마스터")
        headers = ["ID", "INCI명", "한글명", "CAS번호", "유형", "설명",
                    "출처", "생성일", "수정일"]
        ws.append(headers)
        style_header(ws, 1, len(headers))

        try:
            items = json.loads(latest_im[-1].read_text(encoding="utf-8"))
            if isinstance(items, list):
                for i, item in enumerate(items):
                    ws.append([
                        item.get("id", ""),
                        item.get("inci_name", ""),
                        item.get("korean_name", ""),
                        item.get("cas_number", ""),
                        item.get("ingredient_type", ""),
                        str(item.get("description", ""))[:200],
                        item.get("source", ""),
                        str(item.get("created_at", ""))[:19],
                        str(item.get("updated_at", ""))[:19],
                    ])
                    style_data_row(ws, i + 2, 9, is_alt=(i % 2 == 1))
        except Exception as e:
            print(f"  원료마스터 오류: {e}")

        auto_width(ws)

    # ─── 규제정보 시트 ─────────────────────────────────────────────────────────
    latest_reg = sorted(DB_DIR.glob("regulation_cache_*.json"), key=os.path.getmtime)
    if latest_reg:
        ws = wb.create_sheet("규제정보")
        headers = ["출처", "성분명", "INCI명", "최대농도", "제한사항", "수정일"]
        ws.append(headers)
        style_header(ws, 1, len(headers))

        try:
            items = json.loads(latest_reg[-1].read_text(encoding="utf-8"))
            if isinstance(items, list):
                for i, item in enumerate(items):
                    ws.append([
                        item.get("source", ""),
                        item.get("ingredient", ""),
                        item.get("inci_name", ""),
                        item.get("max_concentration", ""),
                        str(item.get("restriction", ""))[:200],
                        str(item.get("updated_at", ""))[:19],
                    ])
                    style_data_row(ws, i + 2, 6, is_alt=(i % 2 == 1))
        except Exception as e:
            print(f"  규제정보 오류: {e}")

        auto_width(ws)

    # ─── 지식베이스 시트 ───────────────────────────────────────────────────────
    latest_kb = sorted(DB_DIR.glob("coching_knowledge_base_*.json"), key=os.path.getmtime)
    if latest_kb:
        ws = wb.create_sheet("지식베이스")
        headers = ["ID", "카테고리", "검색키", "EWG등급", "INCI명",
                    "최대농도", "한국규제", "EU규제", "안전성노트", "버전"]
        ws.append(headers)
        style_header(ws, 1, len(headers))

        try:
            items = json.loads(latest_kb[-1].read_text(encoding="utf-8"))
            if isinstance(items, list):
                for i, item in enumerate(items):
                    d = item.get("data", {})
                    if isinstance(d, str):
                        try:
                            d = json.loads(d)
                        except Exception:
                            d = {}
                    ws.append([
                        item.get("id", ""),
                        item.get("category", ""),
                        item.get("search_key", ""),
                        d.get("ewg_score", ""),
                        d.get("inci_name", ""),
                        d.get("max_concentration", ""),
                        str(d.get("kr_regulation", ""))[:150],
                        str(d.get("eu_regulation", ""))[:150],
                        str(d.get("safety_notes", ""))[:150],
                        item.get("version", ""),
                    ])
                    style_data_row(ws, i + 2, 10, is_alt=(i % 2 == 1))
        except Exception as e:
            print(f"  지식베이스 오류: {e}")

        auto_width(ws)

    # ─── 제품 마스터 시트 ──────────────────────────────────────────────────────
    latest_pm = sorted(DB_DIR.glob("product_master_*.json"), key=os.path.getmtime)
    if latest_pm:
        ws = wb.create_sheet("제품마스터")
        headers = ["ID", "제품명", "브랜드", "카테고리", "출처", "수정일"]
        ws.append(headers)
        style_header(ws, 1, len(headers))

        try:
            items = json.loads(latest_pm[-1].read_text(encoding="utf-8"))
            if isinstance(items, list):
                for i, item in enumerate(items):
                    ws.append([
                        item.get("id", ""),
                        item.get("product_name", item.get("name", "")),
                        item.get("brand", ""),
                        item.get("category", ""),
                        item.get("source", ""),
                        str(item.get("updated_at", ""))[:19],
                    ])
                    style_data_row(ws, i + 2, 6, is_alt=(i % 2 == 1))
        except Exception as e:
            print(f"  제품마스터 오류: {e}")

        auto_width(ws)

    # ─── 화장품업체 시트 ───────────────────────────────────────────────────────
    latest_cc = sorted(DB_DIR.glob("cosmetics_company_*.json"), key=os.path.getmtime)
    if latest_cc:
        ws = wb.create_sheet("화장품업체")
        headers = ["ID", "업체명", "업종", "대표자", "주소", "등록일", "출처"]
        ws.append(headers)
        style_header(ws, 1, len(headers))

        try:
            items = json.loads(latest_cc[-1].read_text(encoding="utf-8"))
            if isinstance(items, list):
                for i, item in enumerate(items):
                    ws.append([
                        item.get("id", ""),
                        item.get("company_name", ""),
                        item.get("business_type", ""),
                        item.get("representative", ""),
                        str(item.get("address", ""))[:100],
                        item.get("registration_date", ""),
                        item.get("source", ""),
                    ])
                    style_data_row(ws, i + 2, 7, is_alt=(i % 2 == 1))
        except Exception as e:
            print(f"  화장품업체 오류: {e}")

        auto_width(ws)

    if wb.sheetnames:
        wb.save(excel_path)
        print(f"  원료DB Excel: {len(wb.sheetnames)}개 시트 → {excel_path.name}")
    else:
        print("  원료DB: 데이터 없음")


# ═══════════════════════════════════════════════════════════════════════════════
# 3. 수집 이력 → collection_log.xlsx
# ═══════════════════════════════════════════════════════════════════════════════
def export_collection_log():
    excel_path = EXCEL_DIR / "COCHING_수집이력.xlsx"

    # 누적 저장 — 기존 파일에서 마지막 타임스탬프 확인
    existing_ts = set()
    if excel_path.exists():
        wb = load_workbook(excel_path)
        if "수집이력" in wb.sheetnames:
            ws = wb["수집이력"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    existing_ts.add(str(row[0]))
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "수집이력" not in wb.sheetnames:
        ws = wb.create_sheet("수집이력")
        headers = ["백업일시", "원료수", "제품수", "규제수", "지식베이스수",
                    "업체수", "백업위치"]
        ws.append(headers)
        style_header(ws, 1, len(headers))
    else:
        ws = wb["수집이력"]

    # backup_summary JSON에서 이력 추출
    new_count = 0
    for f in sorted(DB_DIR.glob("backup_summary_*.json")):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            ts = data.get("timestamp", "")
            if ts in existing_ts:
                continue

            stats_text = data.get("stats", "")
            stats = {}
            for line in stats_text.split("\n"):
                if ":" in line:
                    k, v = line.split(":", 1)
                    stats[k.strip()] = v.strip()

            row_idx = ws.max_row + 1
            ws.append([
                ts,
                stats.get("ingredient_master", ""),
                stats.get("product_master", ""),
                stats.get("regulation_cache", ""),
                stats.get("knowledge_base", ""),
                stats.get("cosmetics_company", ""),
                ", ".join(data.get("backup_locations", [])),
            ])
            style_data_row(ws, row_idx, 7, is_alt=(new_count % 2 == 1))
            new_count += 1
            existing_ts.add(ts)
        except Exception:
            continue

    auto_width(ws)
    wb.save(excel_path)
    print(f"  수집이력 Excel: {new_count}건 추가 (총 {ws.max_row - 1}건) → {excel_path.name}")


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"=== COCHING Excel 누적 저장 [{now}] ===")

    print("\n[1/3] 가이드처방 Excel 변환")
    export_formulations()

    print("\n[2/3] 원료DB Excel 변환")
    export_db_data()

    print("\n[3/3] 수집이력 Excel 변환")
    export_collection_log()

    print(f"\n=== 완료 ===")


if __name__ == "__main__":
    main()
