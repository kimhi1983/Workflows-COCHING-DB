#!/usr/bin/env python3
"""
COCHING 워크플로우 일일 리포트 생성기
- 매일 08:00 크론 실행
- n8n 실행 이력 + PostgreSQL DB 현황 → Excel 저장
- 저장 경로: /mnt/e/COCHING/워크플로우 작업내용/
"""

import sqlite3
import json
import subprocess
import os
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    exit(1)

# ===== 설정 =====
N8N_DB = "/home/kpros/.n8n/database.sqlite"
PG_CONN = {
    "host": "127.0.0.1",
    "port": "5432",
    "db": "coching_db",
    "user": "coching_user",
    "pw": "coching2026!"
}
OUTPUT_DIR = Path("/mnt/e/COCHING/워크플로우 작업내용")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ===== 스타일 =====
HEADER_FONT = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TITLE_FONT = Font(name="맑은 고딕", bold=True, size=14)
NORMAL_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def apply_header(ws, row, cols):
    """헤더 행 스타일 적용"""
    for col_idx in range(1, cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def apply_cell_style(cell, is_data=True):
    """일반 셀 스타일"""
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    if is_data:
        cell.alignment = Alignment(vertical="center")


def auto_width(ws, min_width=10, max_width=50):
    """컬럼 폭 자동 조정"""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value:
                # 한글은 2배 폭
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                lengths.append(length)
        if lengths:
            width = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = width


def pg_query(sql):
    """PostgreSQL 쿼리 실행"""
    pgpass = "/tmp/pgpass_report"
    with open(pgpass, "w") as f:
        f.write(f"{PG_CONN['host']}:{PG_CONN['port']}:{PG_CONN['db']}:{PG_CONN['user']}:{PG_CONN['pw']}")
    os.chmod(pgpass, 0o600)

    cmd = [
        "psql", "-h", PG_CONN["host"], "-U", PG_CONN["user"],
        "-d", PG_CONN["db"], "-t", "-A", "-F", "|", "-c", sql
    ]
    env = os.environ.copy()
    env["PGPASSFILE"] = pgpass
    result = subprocess.run(cmd, capture_output=True, text=True, env=env, timeout=10)

    rows = []
    for line in result.stdout.strip().split("\n"):
        line = line.strip()
        if line:
            rows.append(line.split("|"))
    return rows


def get_n8n_executions():
    """n8n 실행 이력 조회"""
    db = sqlite3.connect(N8N_DB)
    rows = db.execute("""
        SELECT
            e.id,
            w.name,
            e.status,
            e.mode,
            e.startedAt,
            e.stoppedAt,
            e.workflowId
        FROM execution_entity e
        LEFT JOIN workflow_entity w ON e.workflowId = w.id
        ORDER BY e.startedAt DESC
    """).fetchall()
    db.close()
    return rows


def get_workflow_list():
    """n8n 워크플로우 목록"""
    db = sqlite3.connect(N8N_DB)
    rows = db.execute("""
        SELECT id, name, active, createdAt, updatedAt
        FROM workflow_entity
        ORDER BY name
    """).fetchall()
    db.close()
    return rows


def get_db_stats():
    """PostgreSQL 테이블별 데이터 현황"""
    sql = """
    SELECT relname, n_live_tup
    FROM pg_stat_user_tables
    ORDER BY n_live_tup DESC;
    """
    return pg_query(sql)


def get_guide_cache_stats():
    """가이드 캐시 상세 현황"""
    sql = """
    SELECT product_type, skin_type, formula_name, total_wt_percent, wt_valid,
           created_at::text, updated_at::text
    FROM guide_cache
    ORDER BY created_at DESC;
    """
    return pg_query(sql)


def get_guide_coverage():
    """가이드 커버리지 현황"""
    sql = """
    SELECT category, product_type, total_skins, completed, pending
    FROM guide_coverage
    ORDER BY category, product_type;
    """
    return pg_query(sql)


def main():
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    date_display = now.strftime("%Y-%m-%d %H:%M")

    wb = Workbook()

    # ===== Sheet 1: 종합 요약 =====
    ws1 = wb.active
    ws1.title = "종합 요약"

    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"COCHING 워크플로우 일일 리포트 — {date_display}"
    ws1["A1"].font = TITLE_FONT

    # 워크플로우 목록
    ws1["A3"] = "■ 워크플로우 현황"
    ws1["A3"].font = Font(name="맑은 고딕", bold=True, size=12)

    headers = ["워크플로우명", "상태", "워크플로우 ID", "생성일", "최종 수정일"]
    for idx, h in enumerate(headers, 1):
        ws1.cell(row=4, column=idx, value=h)
    apply_header(ws1, 4, len(headers))

    workflows = get_workflow_list()
    for r_idx, wf in enumerate(workflows, 5):
        ws1.cell(row=r_idx, column=1, value=wf[1])
        status = "활성" if wf[2] else "비활성"
        cell = ws1.cell(row=r_idx, column=2, value=status)
        cell.fill = SUCCESS_FILL if wf[2] else ERROR_FILL
        ws1.cell(row=r_idx, column=3, value=wf[0])
        ws1.cell(row=r_idx, column=4, value=wf[3])
        ws1.cell(row=r_idx, column=5, value=wf[4])
        for c in range(1, 6):
            apply_cell_style(ws1.cell(row=r_idx, column=c))

    # DB 현황
    db_start = len(workflows) + 7
    ws1.cell(row=db_start, column=1, value="■ PostgreSQL DB 테이블 현황")
    ws1.cell(row=db_start, column=1).font = Font(name="맑은 고딕", bold=True, size=12)

    db_headers = ["테이블명", "데이터 수"]
    for idx, h in enumerate(db_headers, 1):
        ws1.cell(row=db_start + 1, column=idx, value=h)
    apply_header(ws1, db_start + 1, len(db_headers))

    db_stats = get_db_stats()
    total_records = 0
    for r_idx, stat in enumerate(db_stats, db_start + 2):
        if len(stat) >= 2:
            ws1.cell(row=r_idx, column=1, value=stat[0])
            count = int(stat[1]) if stat[1].strip().isdigit() else 0
            ws1.cell(row=r_idx, column=2, value=count)
            total_records += count
            for c in range(1, 3):
                apply_cell_style(ws1.cell(row=r_idx, column=c))

    total_row = db_start + 2 + len(db_stats)
    ws1.cell(row=total_row, column=1, value="합계")
    ws1.cell(row=total_row, column=1).font = Font(name="맑은 고딕", bold=True, size=10)
    ws1.cell(row=total_row, column=2, value=total_records)
    ws1.cell(row=total_row, column=2).font = Font(name="맑은 고딕", bold=True, size=10)

    auto_width(ws1)

    # ===== Sheet 2: 워크플로우별 실행 이력 =====
    ws2 = wb.create_sheet("실행 이력")

    ws2.merge_cells("A1:H1")
    ws2["A1"] = f"워크플로우 실행 이력 — {date_display}"
    ws2["A1"].font = TITLE_FONT

    exec_headers = ["#", "워크플로우", "상태", "실행 모드", "시작 시간", "종료 시간", "소요 시간(초)", "워크플로우 ID"]
    for idx, h in enumerate(exec_headers, 1):
        ws2.cell(row=2, column=idx, value=h)
    apply_header(ws2, 2, len(exec_headers))

    executions = get_n8n_executions()
    for r_idx, ex in enumerate(executions, 3):
        ws2.cell(row=r_idx, column=1, value=ex[0])  # id
        ws2.cell(row=r_idx, column=2, value=ex[1] or "알 수 없음")  # workflow name
        status_cell = ws2.cell(row=r_idx, column=3, value=ex[2])  # status
        if ex[2] == "success":
            status_cell.fill = SUCCESS_FILL
        elif ex[2] == "error":
            status_cell.fill = ERROR_FILL

        mode_kr = {"manual": "수동", "trigger": "자동(트리거)", "webhook": "웹훅"}.get(ex[3], ex[3])
        ws2.cell(row=r_idx, column=4, value=mode_kr)
        ws2.cell(row=r_idx, column=5, value=ex[4])  # startedAt
        ws2.cell(row=r_idx, column=6, value=ex[5])  # stoppedAt

        # 소요 시간 계산
        duration = ""
        if ex[4] and ex[5]:
            try:
                fmt = "%Y-%m-%d %H:%M:%S.%f"
                start = datetime.strptime(ex[4][:26], fmt)
                stop = datetime.strptime(ex[5][:26], fmt)
                dur = (stop - start).total_seconds()
                duration = round(dur, 1)
            except:
                duration = ""
        ws2.cell(row=r_idx, column=7, value=duration)
        ws2.cell(row=r_idx, column=8, value=ex[6])  # workflowId

        for c in range(1, 9):
            apply_cell_style(ws2.cell(row=r_idx, column=c))

    auto_width(ws2)

    # ===== Sheet 3: 워크플로우별 통계 =====
    ws3 = wb.create_sheet("워크플로우별 통계")

    ws3.merge_cells("A1:G1")
    ws3["A1"] = f"워크플로우별 실행 통계 — {date_display}"
    ws3["A1"].font = TITLE_FONT

    stat_headers = ["워크플로우", "총 실행", "성공", "실패", "성공률(%)", "최근 실행", "평균 소요(초)"]
    for idx, h in enumerate(stat_headers, 1):
        ws3.cell(row=2, column=idx, value=h)
    apply_header(ws3, 2, len(stat_headers))

    # 워크플로우별 집계
    wf_stats = {}
    for ex in executions:
        wf_name = ex[1] or "알 수 없음"
        if wf_name not in wf_stats:
            wf_stats[wf_name] = {"total": 0, "success": 0, "error": 0, "latest": "", "durations": []}
        wf_stats[wf_name]["total"] += 1
        if ex[2] == "success":
            wf_stats[wf_name]["success"] += 1
        elif ex[2] == "error":
            wf_stats[wf_name]["error"] += 1
        if not wf_stats[wf_name]["latest"]:
            wf_stats[wf_name]["latest"] = ex[4] or ""

        if ex[4] and ex[5]:
            try:
                fmt = "%Y-%m-%d %H:%M:%S.%f"
                start = datetime.strptime(ex[4][:26], fmt)
                stop = datetime.strptime(ex[5][:26], fmt)
                dur = (stop - start).total_seconds()
                if dur > 0:
                    wf_stats[wf_name]["durations"].append(dur)
            except:
                pass

    row = 3
    for wf_name, stats in sorted(wf_stats.items()):
        ws3.cell(row=row, column=1, value=wf_name)
        ws3.cell(row=row, column=2, value=stats["total"])
        ws3.cell(row=row, column=3, value=stats["success"])
        ws3.cell(row=row, column=4, value=stats["error"])
        rate = round(stats["success"] / stats["total"] * 100, 1) if stats["total"] > 0 else 0
        ws3.cell(row=row, column=5, value=rate)
        ws3.cell(row=row, column=6, value=stats["latest"])
        avg_dur = round(sum(stats["durations"]) / len(stats["durations"]), 1) if stats["durations"] else ""
        ws3.cell(row=row, column=7, value=avg_dur)

        for c in range(1, 8):
            apply_cell_style(ws3.cell(row=row, column=c))
        row += 1

    auto_width(ws3)

    # ===== Sheet 4: 가이드처방 생성 현황 =====
    ws4 = wb.create_sheet("가이드처방 현황")

    ws4.merge_cells("A1:G1")
    ws4["A1"] = f"가이드처방 생성 현황 — {date_display}"
    ws4["A1"].font = TITLE_FONT

    guide_headers = ["제품유형", "피부타입", "처방명", "배합비(%)", "검증", "생성일", "수정일"]
    for idx, h in enumerate(guide_headers, 1):
        ws4.cell(row=2, column=idx, value=h)
    apply_header(ws4, 2, len(guide_headers))

    guide_data = get_guide_cache_stats()
    for r_idx, g in enumerate(guide_data, 3):
        if len(g) >= 7:
            ws4.cell(row=r_idx, column=1, value=g[0])
            ws4.cell(row=r_idx, column=2, value=g[1])
            ws4.cell(row=r_idx, column=3, value=g[2])
            ws4.cell(row=r_idx, column=4, value=g[3])
            valid = "통과" if g[4] == "t" else "실패"
            valid_cell = ws4.cell(row=r_idx, column=5, value=valid)
            valid_cell.fill = SUCCESS_FILL if g[4] == "t" else ERROR_FILL
            ws4.cell(row=r_idx, column=6, value=g[5])
            ws4.cell(row=r_idx, column=7, value=g[6])
            for c in range(1, 8):
                apply_cell_style(ws4.cell(row=r_idx, column=c))

    # 커버리지 요약
    cov_start = len(guide_data) + 5
    ws4.cell(row=cov_start, column=1, value="■ 가이드 커버리지 (43유형 × 5피부 = 215)")
    ws4.cell(row=cov_start, column=1).font = Font(name="맑은 고딕", bold=True, size=12)

    cov_headers = ["카테고리", "제품유형", "전체", "완료", "미완료"]
    for idx, h in enumerate(cov_headers, 1):
        ws4.cell(row=cov_start + 1, column=idx, value=h)
    apply_header(ws4, cov_start + 1, len(cov_headers))

    coverage = get_guide_coverage()
    total_done = 0
    total_all = 0
    for r_idx, cov in enumerate(coverage, cov_start + 2):
        if len(cov) >= 5:
            ws4.cell(row=r_idx, column=1, value=cov[0])
            ws4.cell(row=r_idx, column=2, value=cov[1])
            ws4.cell(row=r_idx, column=3, value=int(cov[2]) if cov[2].strip().isdigit() else 0)
            ws4.cell(row=r_idx, column=4, value=int(cov[3]) if cov[3].strip().isdigit() else 0)
            ws4.cell(row=r_idx, column=5, value=int(cov[4]) if cov[4].strip().isdigit() else 0)
            total_all += int(cov[2]) if cov[2].strip().isdigit() else 0
            total_done += int(cov[3]) if cov[3].strip().isdigit() else 0
            for c in range(1, 6):
                apply_cell_style(ws4.cell(row=r_idx, column=c))

    # 커버리지 합계
    sum_row = cov_start + 2 + len(coverage)
    ws4.cell(row=sum_row, column=1, value="합계")
    ws4.cell(row=sum_row, column=1).font = Font(name="맑은 고딕", bold=True, size=10)
    ws4.cell(row=sum_row, column=3, value=total_all)
    ws4.cell(row=sum_row, column=4, value=total_done)
    ws4.cell(row=sum_row, column=5, value=total_all - total_done)
    pct = round(total_done / total_all * 100, 1) if total_all > 0 else 0
    ws4.cell(row=sum_row, column=6, value=f"진행률: {pct}%")
    ws4.cell(row=sum_row, column=6).font = Font(name="맑은 고딕", bold=True, size=11)

    auto_width(ws4)

    # ===== 저장 =====
    filename = f"COCHING_워크플로우_리포트_{date_str}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(str(filepath))
    print(f"✅ 리포트 생성 완료: {filepath}")
    print(f"   워크플로우: {len(workflows)}개")
    print(f"   실행 이력: {len(executions)}건")
    print(f"   DB 테이블: {len(db_stats)}개 (총 {total_records}건)")
    print(f"   가이드처방: {len(guide_data)}건 (커버리지 {pct}%)")


if __name__ == "__main__":
    main()
